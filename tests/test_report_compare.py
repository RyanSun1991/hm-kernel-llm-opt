"""Tests for tools/windows_relay/report_compare.py (level + named target version)."""

from __future__ import annotations

import json
import shutil
import sys
import tempfile
import unittest
from io import StringIO
from pathlib import Path

import pytest

REPO_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(REPO_ROOT / "tools" / "windows_relay"))

import report_compare  # type: ignore  # noqa: E402

try:
    from openpyxl import Workbook  # type: ignore
except ImportError:  # pragma: no cover — skip the whole module
    Workbook = None  # type: ignore


pytestmark = pytest.mark.skipif(Workbook is None, reason="openpyxl not installed")


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

def _write_hiperf_xlsx(
    path: Path,
    *,
    total: int,
    processes: list[tuple[str | int, str, int]],
    threads: list[tuple] | None = None,
    libs: list[tuple] | None = None,
    functions: list[tuple] | None = None,
) -> None:
    """Create a hiperf-style workbook with four sheets."""
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "process_instructions_info"
    ws1.append(["pid", "processName", "count"])
    ws1.append(["TOTAL", "", total])
    for pid, name, count in processes:
        ws1.append([pid, name, count])

    ws2 = wb.create_sheet("thread_instructions_info")
    ws2.append(["pid", "processName", "processCount", "tid", "threadName", "threadCount"])
    for row in threads or []:
        ws2.append(list(row))

    ws3 = wb.create_sheet("lib_instructions_info")
    ws3.append([
        "pid", "processName", "processCount", "tid", "threadName", "threadCount",
        "fileId", "libName", "libCount",
    ])
    for row in libs or []:
        ws3.append(list(row))

    ws4 = wb.create_sheet("functions_instructions_info")
    ws4.append([
        "pid", "processName", "processCount", "tid", "threadName", "threadCount",
        "fileId", "libName", "libCount",
        "functionId", "functionName", "functionCount_self", "functionCount_total",
    ])
    for row in functions or []:
        ws4.append(list(row))

    wb.save(path)


def _make_report_tree(
    root: Path,
    *,
    case: str,
    round_n: int,
    step: int,
    total: int,
    processes: list[tuple[str | int, str, int]],
    **extra,
) -> Path:
    case_dir = root / f"PerfLoad_{case}_round{round_n}" / "hiperf" / f"step{step}"
    xlsx = case_dir / f"PerfLoad_{case}_round{round_n}_step{step}_nosym_hiperfReport.xlsx"
    _write_hiperf_xlsx(xlsx, total=total, processes=processes, **extra)
    return xlsx


class _TmpBase(unittest.TestCase):
    def setUp(self) -> None:
        self._root = Path(tempfile.mkdtemp(prefix="hmopt_hiperf_cmp_"))
        self.addCleanup(shutil.rmtree, self._root, ignore_errors=True)

    def tmpdir(self, name: str) -> Path:
        d = self._root / name
        d.mkdir(parents=True, exist_ok=True)
        return d


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

class ToIntTests(unittest.TestCase):
    def test_int(self) -> None:
        self.assertEqual(report_compare._to_int(1000), 1000)

    def test_float_scientific(self) -> None:
        self.assertEqual(report_compare._to_int(1.01886e11), 101_886_000_000)

    def test_comma_string(self) -> None:
        self.assertEqual(report_compare._to_int("1,234,567"), 1_234_567)

    def test_nan_rejected(self) -> None:
        self.assertIsNone(report_compare._to_int(float("nan")))

    def test_empty_rejected(self) -> None:
        self.assertIsNone(report_compare._to_int(""))
        self.assertIsNone(report_compare._to_int(None))


# ---------------------------------------------------------------------------
# find_reports
# ---------------------------------------------------------------------------

class FindReportsTests(_TmpBase):
    def test_walks_report_dir(self) -> None:
        d = self.tmpdir("reports")
        _make_report_tree(d, case="bilibili_0020", round_n=0, step=1, total=10, processes=[(1, "init", 10)])
        _make_report_tree(d, case="bilibili_0020", round_n=0, step=2, total=20, processes=[(1, "init", 20)])
        _make_report_tree(d, case="bilibili_0020", round_n=1, step=1, total=30, processes=[(1, "init", 30)])
        found = report_compare.find_reports(str(d))
        triples = [(r["case"], r["round"], r["step"]) for r in found]
        self.assertEqual(
            triples,
            [("bilibili_0020", 0, 1), ("bilibili_0020", 0, 2), ("bilibili_0020", 1, 1)],
        )

    def test_missing_directory_returns_empty(self) -> None:
        self.assertEqual(report_compare.find_reports(str(self._root / "nope")), [])

    def test_case_name_with_underscores_parsed_correctly(self) -> None:
        d = self.tmpdir("reports2")
        _make_report_tree(d, case="my_complex_case_0042", round_n=3, step=5, total=1, processes=[(1, "x", 1)])
        [entry] = report_compare.find_reports(str(d))
        self.assertEqual(entry["case"], "my_complex_case_0042")
        self.assertEqual(entry["round"], 3)
        self.assertEqual(entry["step"], 5)


# ---------------------------------------------------------------------------
# _validate_target
# ---------------------------------------------------------------------------

class ValidateTargetTests(unittest.TestCase):
    def test_total_needs_no_names(self) -> None:
        report_compare._validate_target("total", process=None, thread=None, lib=None, function=None)

    def test_process_requires_process_name(self) -> None:
        with self.assertRaisesRegex(ValueError, "--process"):
            report_compare._validate_target("process", process=None, thread=None, lib=None, function=None)
        report_compare._validate_target("process", process="p", thread=None, lib=None, function=None)

    def test_thread_requires_process_and_thread(self) -> None:
        with self.assertRaisesRegex(ValueError, "--process.*--thread|--thread.*--process"):
            report_compare._validate_target("thread", process=None, thread=None, lib=None, function=None)
        with self.assertRaisesRegex(ValueError, "--thread"):
            report_compare._validate_target("thread", process="p", thread=None, lib=None, function=None)
        report_compare._validate_target("thread", process="p", thread="t", lib=None, function=None)

    def test_function_requires_all_four(self) -> None:
        with self.assertRaisesRegex(ValueError, "--function"):
            report_compare._validate_target(
                "function", process="p", thread="t", lib="l", function=None,
            )
        report_compare._validate_target(
            "function", process="p", thread="t", lib="l", function="f",
        )

    def test_invalid_level(self) -> None:
        with self.assertRaisesRegex(ValueError, "invalid level"):
            report_compare._validate_target("weird", process=None, thread=None, lib=None, function=None)


# ---------------------------------------------------------------------------
# extract_target_count
# ---------------------------------------------------------------------------

class ExtractTargetCountTests(_TmpBase):
    def _xlsx(self) -> Path:
        path = self._root / "w.xlsx"
        _write_hiperf_xlsx(
            path,
            total=101_886_000_000,
            processes=[
                (0, "swapper", 4_069_975_650),
                (1, "init", 3_417_038),
                (2, "[sysmgr-main]", 2_926_131_503),
                (10, "samgr", 12),
                (11, "samgr", 18),  # same name, different pid — should aggregate
            ],
            threads=[
                (2, "[sysmgr-main]", 2_926_131_503, 2, "sysmgr", 4_121_993),
                (2, "[sysmgr-main]", 2_926_131_503, 8, "sysmgr-reclaim0", 575_817),
                (2, "[sysmgr-main]", 2_926_131_503, 19, "sysmgr-psi-mem", 853_882),
            ],
            libs=[
                (1, "init", 3_417_038, 1, "/bin/init", 3_417_038,
                 2, "/system/lib/ld-musl-aarch64.so.1", 1_259_886),
                (1, "init", 3_417_038, 1, "/bin/init", 3_417_038,
                 411, "sysmgr.elf", 1_241_759),
            ],
            functions=[
                (1, "init", 3_417_038, 1, "/bin/init", 3_417_038,
                 2, "/system/lib/ld-musl-aarch64.so.1", 1_259_886,
                 182, "strlen", 315_638, 315_638),
                (1, "init", 3_417_038, 1, "/bin/init", 3_417_038,
                 2, "/system/lib/ld-musl-aarch64.so.1", 1_259_886,
                 183, "strncmp", 351_994, 351_994),
            ],
        )
        return path

    def test_level_total(self) -> None:
        out = report_compare.extract_target_count(str(self._xlsx()), level="total")
        self.assertEqual(out["value"], 101_886_000_000)
        self.assertTrue(out["found"])
        self.assertEqual(out["total"], 101_886_000_000)

    def test_level_process_found(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="process", process="init",
        )
        self.assertEqual(out["value"], 3_417_038)
        self.assertTrue(out["found"])

    def test_level_process_missing_reports_zero_and_unfound(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="process", process="nonexistent",
        )
        self.assertEqual(out["value"], 0)
        self.assertFalse(out["found"])

    def test_level_process_aggregates_same_name(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="process", process="samgr",
        )
        self.assertEqual(out["value"], 30)  # 12 + 18
        self.assertTrue(out["found"])

    def test_level_thread(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="thread",
            process="[sysmgr-main]", thread="sysmgr-reclaim0",
        )
        self.assertEqual(out["value"], 575_817)
        self.assertTrue(out["found"])

    def test_level_thread_wrong_process(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="thread",
            process="init", thread="sysmgr-reclaim0",
        )
        self.assertFalse(out["found"])
        self.assertEqual(out["value"], 0)

    def test_level_lib(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="lib",
            process="init", thread="/bin/init", lib="sysmgr.elf",
        )
        self.assertEqual(out["value"], 1_241_759)
        self.assertTrue(out["found"])

    def test_level_function(self) -> None:
        out = report_compare.extract_target_count(
            str(self._xlsx()), level="function",
            process="init", thread="/bin/init",
            lib="/system/lib/ld-musl-aarch64.so.1", function="strncmp",
        )
        self.assertEqual(out["value"], 351_994)
        self.assertTrue(out["found"])

    def test_missing_required_name_raises(self) -> None:
        with self.assertRaisesRegex(ValueError, "--thread"):
            report_compare.extract_target_count(
                str(self._xlsx()), level="thread", process="init",
            )


# ---------------------------------------------------------------------------
# compare_reports
# ---------------------------------------------------------------------------

class CompareReportsTests(_TmpBase):
    def test_level_total_happy_path(self) -> None:
        base = self.tmpdir("base")
        cand = self.tmpdir("cand")
        _make_report_tree(base, case="bilibili_0020", round_n=0, step=1, total=1000, processes=[(1, "p", 1000)])
        _make_report_tree(cand, case="bilibili_0020", round_n=0, step=1, total=900, processes=[(1, "p", 900)])

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand), level="total",
        )
        self.assertTrue(result["success"])
        self.assertEqual(result["level"], "total")
        self.assertEqual(result["target"], {})
        self.assertEqual(result["aggregate"]["baseline"], 1000)
        self.assertEqual(result["aggregate"]["candidate"], 900)
        self.assertEqual(result["aggregate"]["delta"], -100)
        self.assertEqual(result["aggregate"]["delta_pct"], -10.0)
        [pair] = result["reports"]
        self.assertEqual(pair["baseline"], 1000)
        self.assertEqual(pair["candidate"], 900)
        self.assertEqual(pair["delta"], -100)

    def test_level_process(self) -> None:
        base = self.tmpdir("base_p")
        cand = self.tmpdir("cand_p")
        _make_report_tree(
            base, case="a", round_n=0, step=1, total=1000,
            processes=[(1, "samgr", 400), (2, "init", 600)],
        )
        _make_report_tree(
            cand, case="a", round_n=0, step=1, total=950,
            processes=[(1, "samgr", 380), (2, "init", 570)],
        )

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand),
            level="process", process="samgr",
        )
        self.assertTrue(result["success"])
        self.assertEqual(result["target"], {"process": "samgr"})
        self.assertEqual(result["aggregate"]["baseline"], 400)
        self.assertEqual(result["aggregate"]["candidate"], 380)
        self.assertEqual(result["aggregate"]["delta"], -20)

    def test_level_thread(self) -> None:
        base = self.tmpdir("base_t")
        cand = self.tmpdir("cand_t")
        common = dict(
            processes=[(2, "[sysmgr-main]", 1000)],
            threads=[
                (2, "[sysmgr-main]", 1000, 8, "sysmgr-reclaim0", 400),
                (2, "[sysmgr-main]", 1000, 19, "sysmgr-psi-mem", 600),
            ],
        )
        _make_report_tree(base, case="a", round_n=0, step=1, total=1000, **common)
        _make_report_tree(
            cand, case="a", round_n=0, step=1, total=900,
            processes=[(2, "[sysmgr-main]", 900)],
            threads=[
                (2, "[sysmgr-main]", 900, 8, "sysmgr-reclaim0", 320),
                (2, "[sysmgr-main]", 900, 19, "sysmgr-psi-mem", 580),
            ],
        )

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand),
            level="thread", process="[sysmgr-main]", thread="sysmgr-reclaim0",
        )
        self.assertTrue(result["success"])
        self.assertEqual(result["target"], {"process": "[sysmgr-main]", "thread": "sysmgr-reclaim0"})
        self.assertEqual(result["aggregate"]["baseline"], 400)
        self.assertEqual(result["aggregate"]["candidate"], 320)
        self.assertEqual(result["aggregate"]["delta"], -80)

    def test_level_function(self) -> None:
        base = self.tmpdir("base_f")
        cand = self.tmpdir("cand_f")
        common_kwargs = dict(
            processes=[(1, "init", 100)],
            threads=[(1, "init", 100, 1, "/bin/init", 100)],
            libs=[(1, "init", 100, 1, "/bin/init", 100, 2, "ld-musl.so", 100)],
        )
        _make_report_tree(
            base, case="a", round_n=0, step=1, total=100,
            functions=[
                (1, "init", 100, 1, "/bin/init", 100, 2, "ld-musl.so", 100, 1, "foo", 0, 60),
                (1, "init", 100, 1, "/bin/init", 100, 2, "ld-musl.so", 100, 2, "bar", 0, 40),
            ],
            **common_kwargs,
        )
        _make_report_tree(
            cand, case="a", round_n=0, step=1, total=80,
            functions=[
                (1, "init", 100, 1, "/bin/init", 100, 2, "ld-musl.so", 100, 1, "foo", 0, 30),
                (1, "init", 100, 1, "/bin/init", 100, 2, "ld-musl.so", 100, 2, "bar", 0, 50),
            ],
            **common_kwargs,
        )

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand),
            level="function",
            process="init", thread="/bin/init", lib="ld-musl.so", function="foo",
        )
        self.assertEqual(result["aggregate"]["baseline"], 60)
        self.assertEqual(result["aggregate"]["candidate"], 30)
        self.assertEqual(result["aggregate"]["delta"], -30)

    def test_multiple_pairs_sum_to_aggregate(self) -> None:
        base = self.tmpdir("base_m")
        cand = self.tmpdir("cand_m")
        _make_report_tree(base, case="a", round_n=0, step=1, total=10, processes=[(1, "p", 10)])
        _make_report_tree(base, case="b", round_n=0, step=1, total=20, processes=[(1, "p", 20)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=12, processes=[(1, "p", 12)])
        _make_report_tree(cand, case="b", round_n=0, step=1, total=15, processes=[(1, "p", 15)])

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand), level="total",
        )
        self.assertTrue(result["success"])
        self.assertEqual(result["aggregate"]["baseline"], 30)
        self.assertEqual(result["aggregate"]["candidate"], 27)
        self.assertEqual(result["aggregate"]["delta"], -3)
        self.assertEqual(result["aggregate"]["pairs_compared"], 2)

    def test_missing_on_one_side_flagged(self) -> None:
        base = self.tmpdir("base_miss")
        cand = self.tmpdir("cand_miss")
        _make_report_tree(base, case="a", round_n=0, step=1, total=10, processes=[(1, "p", 10)])
        _make_report_tree(base, case="b", round_n=0, step=1, total=10, processes=[(1, "p", 10)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=11, processes=[(1, "p", 11)])

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand), level="total",
        )
        self.assertFalse(result["success"])
        self.assertEqual(result["aggregate"]["pairs_missing_candidate"], 1)
        by_case = {r["case"]: r for r in result["reports"]}
        self.assertEqual(by_case["b"]["missing"], "candidate")

    def test_target_not_found_reports_zero_and_unfound(self) -> None:
        base = self.tmpdir("base_nf")
        cand = self.tmpdir("cand_nf")
        _make_report_tree(base, case="a", round_n=0, step=1, total=100, processes=[(1, "samgr", 100)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=100, processes=[(1, "init", 100)])

        result = report_compare.compare_reports(
            baseline_dir=str(base), candidate_dir=str(cand),
            level="process", process="samgr",
        )
        self.assertTrue(result["success"])  # pairing worked, target simply absent on cand
        [pair] = result["reports"]
        self.assertEqual(pair["baseline"], 100)
        self.assertTrue(pair["baseline_found"])
        self.assertEqual(pair["candidate"], 0)
        self.assertFalse(pair["candidate_found"])
        self.assertEqual(pair["delta"], -100)

    def test_validates_required_names(self) -> None:
        base = self.tmpdir("base_v")
        cand = self.tmpdir("cand_v")
        _make_report_tree(base, case="a", round_n=0, step=1, total=10, processes=[(1, "p", 10)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=12, processes=[(1, "p", 12)])
        with self.assertRaisesRegex(ValueError, "--process"):
            report_compare.compare_reports(
                baseline_dir=str(base), candidate_dir=str(cand), level="process",
            )

    def test_cli_main_emits_json(self) -> None:
        base = self.tmpdir("cli_base")
        cand = self.tmpdir("cli_cand")
        _make_report_tree(base, case="a", round_n=0, step=1, total=100, processes=[(1, "p", 100)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=120, processes=[(1, "p", 120)])

        buf = StringIO()
        original = sys.stdout
        sys.stdout = buf
        try:
            rc = report_compare.main(
                ["--baseline", str(base), "--candidate", str(cand), "--level", "total"],
            )
        finally:
            sys.stdout = original
        self.assertEqual(rc, 0)
        payload = json.loads(buf.getvalue().strip())
        self.assertTrue(payload["success"])
        self.assertEqual(payload["aggregate"]["delta"], 20)

    def test_cli_main_missing_required_name_emits_error(self) -> None:
        base = self.tmpdir("cli_bad_base")
        cand = self.tmpdir("cli_bad_cand")
        _make_report_tree(base, case="a", round_n=0, step=1, total=10, processes=[(1, "p", 10)])
        _make_report_tree(cand, case="a", round_n=0, step=1, total=12, processes=[(1, "p", 12)])

        buf = StringIO()
        original = sys.stdout
        sys.stdout = buf
        try:
            rc = report_compare.main(
                ["--baseline", str(base), "--candidate", str(cand), "--level", "thread"],
            )
        finally:
            sys.stdout = original
        self.assertEqual(rc, 1)
        payload = json.loads(buf.getvalue().strip())
        self.assertFalse(payload["success"])
        self.assertIn("--process", payload["error"])


if __name__ == "__main__":
    unittest.main()
