# -*- coding: utf-8 -*-
"""Parse lmbench autotest result xlsx into a compact, LLM-friendly digest.

Self-contained (stdlib + openpyxl only) so it runs on the bare Windows relay
*and* is importable platform-side. It reads the two artifacts the
`D:\\LmbenchAutoTest\\` framework produces:

  total_result_Hongmeng_<ts>.xlsx        per-run absolute numbers, one row per
                                          (system=core, tool, metric, command):
                                          average / std / Discrete / value0..N / units
  HM_Linux_lmbench_result_<ts>.xlsx       HM-vs-Linux comparison, per metric:
                                          HM_<core> / linux_<core> / 权重 / 差距% / 得分

and emits a small JSON digest with three comparison lines:
  - hm_vs_linux : weighted gap per core + overall + top regressions/wins
  - vs_previous : current vs the previous timestamped total_result (the patch A/B
                  delta when stock then feature are run back-to-back)
  - anomalies   : metrics whose run-to-run Discrete (dispersion) is high

Direction matters: bandwidth (units like MB/s, `bw_*`) is higher-is-better;
latency (`lat_*`, time units) is lower-is-better. We normalize both so a positive
number always means "better".

CLI:
  python lmbench_digest.py --total <total.xlsx> [--hm-linux <cmp.xlsx>] \
         [--prev <prev_total.xlsx>] [--top 8] [--md]
"""
from __future__ import annotations

import argparse
import json
import math
import re
import sys
from pathlib import Path

try:
    import openpyxl  # type: ignore
except ImportError:  # pragma: no cover - relay auto-installs it
    openpyxl = None

_HIGH_DISCRETE = 0.05  # dispersion above this is flagged as noisy


# --------------------------------------------------------------------------- #
# helpers
# --------------------------------------------------------------------------- #
def _num(v):
    """Best-effort *finite* float from a cell that may be a number or '1.04%'.

    Returns None for blanks, non-numeric text, AND non-finite values (nan/inf):
    a 'nan' gap cell must not propagate into the weighted-gap math, and NaN is
    invalid JSON (RFC 8259) — strict consumers downstream (the MCP/JSON-RPC
    boundary) reject it, which would silently break the agent's status polling.
    """
    if v is None:
        return None
    if isinstance(v, (int, float)):
        f = float(v)
        return f if math.isfinite(f) else None
    s = str(v).strip().replace("%", "").replace(",", "")
    if not s:
        return None
    try:
        f = float(s)
    except ValueError:
        return None
    return f if math.isfinite(f) else None


def _json_safe(obj):
    """Recursively replace any non-finite float (nan/inf) with None so the digest
    is always strict-JSON serializable across the relay -> MCP -> agent boundary."""
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    if isinstance(obj, dict):
        return {k: _json_safe(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_json_safe(x) for x in obj]
    return obj


def _direction(command: str, units: str) -> str:
    """higher_better | lower_better | unknown — from units, then command name."""
    u = (units or "").strip().lower()
    c = (command or "").strip().lower()
    if "/s" in u or "b/s" in u or c.startswith("bw") or "bandwidth" in c:
        return "higher_better"
    if c.startswith("lat") or "latency" in c or any(t in u for t in
            ("sec", "second", "micro", "nano", "milli", "us", "ns", "ms")):
        return "lower_better"
    return "unknown"


def _norm(delta_pct: float, direction: str):
    """Normalize a raw delta% so positive = improvement, given the metric direction."""
    if direction == "higher_better":
        return delta_pct
    if direction == "lower_better":
        return -delta_pct
    return None  # unknown -> cannot sign


def _rows(path, sheet="result"):
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb.worksheets[0]
    it = ws.iter_rows(values_only=True)
    header = [str(h).strip() if h is not None else "" for h in next(it)]
    idx = {h: i for i, h in enumerate(header)}
    out = []
    for r in it:
        if r is None or all(c is None for c in r):
            continue
        out.append((r, idx))
    return out, idx


# --------------------------------------------------------------------------- #
# parsers
# --------------------------------------------------------------------------- #
def parse_total(path):
    """-> list of {system, tool, metric, command, average, std, discrete, units, direction}.

    The framework writes ``average`` as an Excel formula. The file is generated
    programmatically and never opened in Excel, so under ``data_only=True`` that
    cell has no cached value and reads as None — which silently zeroed every A/B
    comparison. When ``average`` is empty we derive it from the raw ``value0..N``
    sample columns (literals), which is exactly what the average represents.
    """
    rows, idx = _rows(path)
    value_cols = sorted((h for h in idx if re.fullmatch(r"value\d+", h)),
                        key=lambda h: int(h[5:]))
    out = []
    for r, _ in rows:
        def g(name):
            return r[idx[name]] if name in idx and idx[name] < len(r) else None
        cmd = str(g("command") or "")
        units = str(g("units") or "")
        avg = _num(g("average"))
        if avg is None:  # empty/uncached formula -> mean of the present sample values
            samples = [s for s in (_num(g(v)) for v in value_cols) if s is not None]
            if samples:
                avg = sum(samples) / len(samples)
        out.append({
            "system": str(g("system") or ""),
            "tool": str(g("tool") or ""),
            "metric": str(g("metric") or ""),
            "command": cmd,
            "average": avg,
            "std": _num(g("standard_deviation")),
            "discrete": _num(g("Discrete")),
            "units": units,
            "direction": _direction(cmd, units),
        })
    return out


def parse_hm_linux(path):
    """-> list of {module, indicator, tool, metric, command, cores: {core: {hm, linux, weight, gap_pct, score}}}."""
    rows, idx = _rows(path)
    cores = []
    for h in idx:
        m = re.match(r"HM_(.+)$", h)
        if m:
            cores.append(m.group(1))
    out = []
    for r, _ in rows:
        def g(name):
            return r[idx[name]] if name in idx and idx[name] < len(r) else None
        rec = {
            "module": str(g("benchmark_module") or ""),
            "indicator": str(g("performance_indicator") or ""),
            "tool": str(g("tool") or ""),
            "metric": str(g("metric") or ""),
            "command": str(g("command") or ""),
            "cores": {},
        }
        for core in cores:
            rec["cores"][core] = {
                "hm": _num(g(f"HM_{core}")),
                "linux": _num(g(f"linux_{core}")),
                "weight": _num(g(f"权重_{core}")),
                "gap_pct": _num(g(f"差距_{core}")),
                "score": _num(g(f"得分_{core}")),
            }
        out.append(rec)
    return out


# --------------------------------------------------------------------------- #
# digest
# --------------------------------------------------------------------------- #
def _hm_vs_linux(rows, top_n):
    by_core = {}
    norm_items = []  # (hm_advantage_pct, record-for-listing)
    for rec in rows:
        direction = _direction(rec["command"], "")
        for core, c in rec["cores"].items():
            gap, w = c.get("gap_pct"), c.get("weight")
            if gap is None:
                continue
            d = by_core.setdefault(core, {"wsum": 0.0, "wgap": 0.0, "score_sum": 0.0})
            if w is not None:
                d["wsum"] += w
                d["wgap"] += w * gap
            if c.get("score") is not None:
                d["score_sum"] += c["score"]
            adv = _norm(gap, direction)
            if adv is not None:
                norm_items.append((adv, {
                    "module": rec["module"], "metric": rec["metric"],
                    "command": rec["command"], "core": core,
                    "hm": c.get("hm"), "linux": c.get("linux"),
                    "gap_pct": round(gap, 2), "weight": w, "direction": direction,
                }))
    per_core = {core: {
        "weighted_gap_pct": round(d["wgap"] / d["wsum"], 3) if d["wsum"] else None,
        "weight_sum": round(d["wsum"], 2), "score_sum": round(d["score_sum"], 2),
    } for core, d in by_core.items()}
    wsum = sum(d["wsum"] for d in by_core.values())
    wgap = sum(d["wgap"] for d in by_core.values())
    norm_items.sort(key=lambda x: x[0])
    return {
        "by_core": per_core,
        "overall_weighted_gap_pct": round(wgap / wsum, 3) if wsum else None,
        "top_regressions": [x[1] for x in norm_items[:top_n]],          # HM worse than Linux
        "top_wins": [x[1] for x in reversed(norm_items[-top_n:])],      # HM beats Linux
    }


def _match_key(r):
    """Normalized (system, tool, metric, command) — strip + collapse inner whitespace
    so trivial formatting differences between two runs don't defeat metric matching."""
    return tuple(re.sub(r"\s+", " ", str(r.get(k) or "")).strip()
                 for k in ("system", "tool", "metric", "command"))


def _vs_previous(cur, prev, top_n):
    pmap = {_match_key(r): r for r in prev}
    items, matched, improved, regressed, unknown = [], 0, 0, 0, 0
    for r in cur:
        p = pmap.get(_match_key(r))
        if not p or not p.get("average") or r.get("average") is None:
            continue
        matched += 1
        delta_pct = (r["average"] - p["average"]) / p["average"] * 100.0
        imp = _norm(delta_pct, r["direction"])
        if imp is None:
            unknown += 1
        elif imp > 0:
            improved += 1
        else:
            regressed += 1
        items.append((imp if imp is not None else 0.0, {
            "system": r["system"], "tool": r["tool"], "metric": r["metric"],
            "command": r["command"], "units": r["units"], "direction": r["direction"],
            "prev_avg": round(p["average"], 2), "cur_avg": round(r["average"], 2),
            "delta_pct": round(delta_pct, 2),
            "improvement_pct": None if imp is None else round(imp, 2),
        }))
    items.sort(key=lambda x: x[0])
    out = {
        "matched": matched, "improved": improved, "regressed": regressed,
        "n_unknown_direction": unknown,
        "top_regressions": [x[1] for x in items[:top_n]],
        "top_improvements": [x[1] for x in reversed(items[-top_n:])],
    }
    if matched == 0:
        out["diagnostic"] = _match_diagnostic(cur, prev)
    return out


def _match_diagnostic(cur, prev):
    """When 0 metrics matched, surface WHY so it can be fixed without re-running the
    2-5h suite: row counts, average-presence, key overlap, and a few sample keys."""
    cur_keys = [_match_key(r) for r in cur]
    prev_keys = {_match_key(r) for r in prev}
    overlap = sum(1 for k in cur_keys if k in prev_keys)
    n_cur_avg = sum(1 for r in cur if r.get("average") is not None)
    n_prev_avg = sum(1 for r in prev if r.get("average") is not None)
    if n_cur_avg == 0 or n_prev_avg == 0:
        reason = ("the 'average' column was not parsed in one/both files (header name "
                  "or sheet mismatch) — check *_headers below")
    elif overlap == 0:
        reason = ("metric keys (system,tool,metric,command) do not overlap between "
                  "the two runs — compare sample_*_keys below")
    else:
        reason = "keys overlap but averages were missing on the overlapping rows"
    return {
        "reason": reason,
        "n_cur": len(cur), "n_prev": len(prev),
        "n_cur_with_avg": n_cur_avg, "n_prev_with_avg": n_prev_avg,
        "n_key_overlap": overlap,
        "sample_cur_keys": [list(k) for k in cur_keys[:4]],
        "sample_prev_keys": [list(k) for k in list(prev_keys)[:4]],
    }


def _header_names(path):
    """Raw first-row header of a total_result sheet — the decisive evidence when the
    'average' column isn't being read (renamed column, wrong sheet, title row)."""
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb["result"] if "result" in wb.sheetnames else wb.worksheets[0]
        row = next(ws.iter_rows(values_only=True))
        return [str(h).strip() if h is not None else "" for h in row]
    except Exception as exc:  # never raise from a diagnostic
        return [f"<header read failed: {type(exc).__name__}: {exc}>"]


def _first_data_row(path):
    """Raw first non-empty data row (data_only) — shows whether average/value cells
    are themselves empty (uncached formulas), the only thing left to check if even
    the value-column fallback yields no numbers."""
    try:
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        ws = wb["result"] if "result" in wb.sheetnames else wb.worksheets[0]
        it = ws.iter_rows(values_only=True)
        next(it, None)  # skip header
        for r in it:
            if r and not all(c is None for c in r):
                return [c if isinstance(c, (int, float)) or c is None else str(c) for c in r]
        return []
    except Exception as exc:  # never raise from a diagnostic
        return [f"<row read failed: {type(exc).__name__}: {exc}>"]


def _is_lockfile(path) -> bool:
    """Excel writes a hidden ~$<name>.xlsx companion while a workbook is open."""
    name = Path(str(path)).name
    return name.startswith("~$") or name.startswith(".~")


def build_digest(total_path, hm_linux_path=None, prev_total_path=None, top_n=8):
    if openpyxl is None:
        return {"ok": False, "error": "openpyxl not installed on this host"}
    try:
        cur = parse_total(total_path)  # core result — a failure here is fatal
    except Exception as exc:  # never raise on the relay
        return {"ok": False, "error": f"total parse failed: {type(exc).__name__}: {exc}"}

    d = {
        "ok": True,
        "files": {
            "total": str(total_path),
            "hm_linux": str(hm_linux_path) if hm_linux_path else None,
            "prev_total": str(prev_total_path) if prev_total_path else None,
        },
        "n_metrics": len(cur),
        "warnings": [],
        "anomalies": sorted(
            [{"system": r["system"], "metric": r["metric"], "command": r["command"],
              "discrete": r["discrete"]}
             for r in cur if (r["discrete"] or 0) >= _HIGH_DISCRETE],
            key=lambda x: -(x["discrete"] or 0))[:top_n],
    }

    # hm_vs_linux and vs_previous are best-effort: a locked / lock-file / corrupt
    # input degrades that one section to a warning rather than killing the digest.
    if hm_linux_path and not _is_lockfile(hm_linux_path) and Path(hm_linux_path).exists():
        try:
            d["hm_vs_linux"] = _hm_vs_linux(parse_hm_linux(hm_linux_path), top_n)
        except Exception as exc:
            d["warnings"].append(f"hm_linux skipped: {type(exc).__name__}: {exc}")
    elif hm_linux_path and _is_lockfile(hm_linux_path):
        d["warnings"].append(f"hm_linux skipped: looks like an Excel lock file ({Path(str(hm_linux_path)).name})")

    if prev_total_path and not _is_lockfile(prev_total_path) and Path(prev_total_path).exists():
        try:
            d["vs_previous"] = _vs_previous(cur, parse_total(prev_total_path), top_n)
            if d["vs_previous"].get("matched") == 0 and "diagnostic" in d["vs_previous"]:
                diag = d["vs_previous"]["diagnostic"]
                diag["cur_headers"] = _header_names(total_path)
                diag["prev_headers"] = _header_names(prev_total_path)
                diag["cur_first_row"] = _first_data_row(total_path)
                diag["prev_first_row"] = _first_data_row(prev_total_path)
        except Exception as exc:
            d["warnings"].append(f"vs_previous skipped: {type(exc).__name__}: {exc}")
    elif prev_total_path and _is_lockfile(prev_total_path):
        d["warnings"].append(f"vs_previous skipped: looks like an Excel lock file ({Path(str(prev_total_path)).name})")

    return _json_safe(d)


def digest_markdown(d):
    if not d.get("ok"):
        return f"lmbench digest unavailable: {d.get('error')}"
    L = [f"# lmbench digest ({d['n_metrics']} metrics)"]
    hv = d.get("hm_vs_linux")
    if hv:
        L.append(f"\n## HM vs Linux — overall weighted gap {hv['overall_weighted_gap_pct']}% "
                 + " ".join(f"[{c} {v['weighted_gap_pct']}%]" for c, v in hv["by_core"].items()))
        L.append("Worst vs Linux (HM slower):")
        for x in hv["top_regressions"][:5]:
            L.append(f"- {x['core']} `{x['command']}` gap {x['gap_pct']}% ({x['module']})")
    vp = d.get("vs_previous")
    if vp:
        L.append(f"\n## vs previous run — improved {vp['improved']} / regressed {vp['regressed']} "
                 f"/ unknown-dir {vp['n_unknown_direction']} (matched {vp['matched']})")
        diag = vp.get("diagnostic")
        if diag:
            L.append(f"**0 metrics matched — {diag['reason']}**")
            L.append(f"- rows: cur={diag['n_cur']} (avg on {diag['n_cur_with_avg']}) "
                     f"prev={diag['n_prev']} (avg on {diag['n_prev_with_avg']}) key-overlap={diag['n_key_overlap']}")
            if diag.get("cur_headers") is not None:
                L.append(f"- cur headers: {diag['cur_headers']}")
                L.append(f"- prev headers: {diag['prev_headers']}")
            for k in diag["sample_cur_keys"][:3]:
                L.append(f"- cur key sample: {k}")
            for k in diag["sample_prev_keys"][:3]:
                L.append(f"- prev key sample: {k}")
        # direction-aware: improvement_pct > 0 means "better" regardless of metric
        # direction; only list rows that actually moved the wrong/right way.
        regs = [x for x in vp["top_regressions"]
                if x.get("improvement_pct") is not None and x["improvement_pct"] < 0]
        imps = [x for x in vp["top_improvements"]
                if x.get("improvement_pct") is not None and x["improvement_pct"] > 0]
        L.append("Top regressions (worse, direction-aware):" + ("" if regs else " none"))
        for x in regs[:5]:
            L.append(f"- {x['system']} `{x['command']}` {x['improvement_pct']}% "
                     f"({x['prev_avg']}→{x['cur_avg']} {x['units']}, raw delta {x['delta_pct']}%)")
        L.append("Top improvements (better, direction-aware):" + ("" if imps else " none"))
        for x in imps[:5]:
            L.append(f"- {x['system']} `{x['command']}` +{x['improvement_pct']}% "
                     f"({x['prev_avg']}→{x['cur_avg']} {x['units']}, raw delta {x['delta_pct']}%)")
    if d.get("anomalies"):
        L.append("\n## High-dispersion metrics (treat with caution): "
                 + ", ".join(f"{a['command']}({a['discrete']})" for a in d["anomalies"][:5]))
    if d.get("warnings"):
        L.append("\n## Warnings")
        for w in d["warnings"]:
            L.append(f"- {w}")
    return "\n".join(L)


def main(argv=None):
    ap = argparse.ArgumentParser()
    ap.add_argument("--total", required=True)
    ap.add_argument("--hm-linux")
    ap.add_argument("--prev")
    ap.add_argument("--top", type=int, default=8)
    ap.add_argument("--md", action="store_true", help="emit markdown instead of JSON")
    a = ap.parse_args(argv)
    d = build_digest(a.total, a.hm_linux, a.prev, top_n=a.top)
    print(digest_markdown(d) if a.md else json.dumps(d, ensure_ascii=False, indent=2))
    return 0 if d.get("ok") else 1


if __name__ == "__main__":
    sys.exit(main())
