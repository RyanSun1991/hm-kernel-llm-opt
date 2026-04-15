"""Compare two modelCase test report directories on instruction count.

The modelCase pipeline produces, for every (case, round, step) triple::

  <report_dir>/
    PerfLoad_<case>_round<N>/
      hiperf/
        step<M>/
          PerfLoad_<case>_round<N>_step<M>_nosym_hiperfReport.xlsx

Each workbook has four worksheets:

  1. ``process_instructions_info``    pid | processName | count
       (first data row pid="TOTAL" holds the overall instruction count)
  2. ``thread_instructions_info``     pid | processName | processCount
                                       | tid | threadName | threadCount
  3. ``lib_instructions_info``        ... | fileId | libName | libCount
  4. ``functions_instructions_info``  ... | functionId | functionName
                                       | functionCount_self | functionCount_total

The comparison target is picked by ``--level`` and the names at that
level and above.  Each level requires just the names of its own and
parent tiers — you don't have to specify deeper names you don't care
about::

  python report_compare.py --baseline B --candidate C --level total
  python report_compare.py --baseline B --candidate C --level process \\
      --process samgr
  python report_compare.py --baseline B --candidate C --level thread \\
      --process sysmgr-main --thread sysmgr-reclaim0
  python report_compare.py --baseline B --candidate C --level lib \\
      --process init --thread /bin/init --lib /system/lib/ld-musl-aarch64.so.1
  python report_compare.py --baseline B --candidate C --level function \\
      --process init --thread /bin/init --lib /system/lib/ld-musl-aarch64.so.1 \\
      --function strlen

For every matched workbook pair the script emits baseline count,
candidate count, delta and delta_pct.  When the target name appears
multiple times (e.g. two pids share a processName) the counts are
summed.  When the name is absent the count is reported as 0 with
``baseline_found`` / ``candidate_found`` set to False.

Requires ``openpyxl`` on the executing host (``pip install openpyxl``).
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from typing import Any, Iterable

try:  # pragma: no cover — availability is what we test, not the import itself
    from openpyxl import load_workbook  # type: ignore
except ImportError:  # pragma: no cover
    load_workbook = None  # type: ignore[assignment]


# ---------------------------------------------------------------------------
# Report-file discovery
# ---------------------------------------------------------------------------

REPORT_FILE_PATTERN = re.compile(
    r"^PerfLoad_(?P<case>.+?)_round(?P<round>\d+)_step(?P<step>\d+)"
    r"_nosym_hiperfReport\.xlsx$",
    re.IGNORECASE,
)


def find_reports(report_dir: str) -> list[dict[str, Any]]:
    """Walk *report_dir* and return every matching hiperf xlsx."""
    if not os.path.isdir(report_dir):
        return []
    found: list[dict[str, Any]] = []
    for root, _dirs, files in os.walk(report_dir):
        for name in files:
            match = REPORT_FILE_PATTERN.match(name)
            if not match:
                continue
            found.append({
                "case": match.group("case"),
                "round": int(match.group("round")),
                "step": int(match.group("step")),
                "path": os.path.join(root, name),
            })
    found.sort(key=lambda r: (r["case"], r["round"], r["step"]))
    return found


# ---------------------------------------------------------------------------
# Value coercion
# ---------------------------------------------------------------------------

def _to_int(value: Any) -> int | None:
    """Best-effort coerce a cell value to int.

    Handles plain ints, floats (Excel often stores integers as floats
    and may display them in scientific notation, e.g. 1.01886e+11), and
    numeric strings with commas or underscores.
    """
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value != value:  # NaN
            return None
        return int(round(value))
    if isinstance(value, str):
        text = value.replace(",", "").replace("_", "").strip()
        if not text:
            return None
        try:
            return int(text)
        except ValueError:
            try:
                return int(round(float(text)))
            except ValueError:
                return None
    return None


def _clean_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


# ---------------------------------------------------------------------------
# Worksheet parsing
# ---------------------------------------------------------------------------

PROCESS_SHEET = "process_instructions_info"
THREAD_SHEET = "thread_instructions_info"
LIB_SHEET = "lib_instructions_info"
FUNCTION_SHEET = "functions_instructions_info"


def _iter_data_rows(ws: Any, min_cols: int) -> Iterable[tuple]:
    """Yield data rows (header skipped) padded to at least *min_cols*."""
    for idx, row in enumerate(ws.iter_rows(values_only=True)):
        if idx == 0:  # header
            continue
        if not row or all(v is None or _clean_str(v) == "" for v in row):
            continue
        padded = tuple(row) + (None,) * max(0, min_cols - len(row))
        yield padded


def _parse_process_sheet(ws: Any) -> tuple[int | None, dict[str, int]]:
    """Return ``(total, {processName: sum_count})``.

    ``total`` is the count on the ``pid="TOTAL"`` row; ``None`` if
    missing.  Process rows are summed by name so two pids with the
    same processName collapse into one entry.
    """
    total: int | None = None
    processes: dict[str, int] = {}
    for row in _iter_data_rows(ws, min_cols=3):
        pid_raw, process_name, count = row[0], row[1], row[2]
        pid = _clean_str(pid_raw)
        if pid.upper() == "TOTAL":
            total = _to_int(count)
            continue
        value = _to_int(count)
        if value is None:
            continue
        key = _clean_str(process_name) or pid or "<unknown>"
        processes[key] = processes.get(key, 0) + value
    return total, processes


def _sum_thread_sheet(ws: Any, *, process: str, thread: str) -> tuple[int, bool]:
    """Return ``(sum_threadCount, found)`` for rows matching ``process``+``thread``."""
    total = 0
    found = False
    for row in _iter_data_rows(ws, min_cols=6):
        _pid, process_name, _pcnt, _tid, thread_name, thread_count = row[:6]
        if _clean_str(process_name) != process or _clean_str(thread_name) != thread:
            continue
        value = _to_int(thread_count)
        if value is None:
            continue
        total += value
        found = True
    return total, found


def _sum_lib_sheet(ws: Any, *, process: str, thread: str, lib: str) -> tuple[int, bool]:
    """Return ``(sum_libCount, found)`` for rows matching process+thread+lib."""
    total = 0
    found = False
    for row in _iter_data_rows(ws, min_cols=9):
        _pid, process_name, _pcnt, _tid, thread_name, _tcnt, _fid, lib_name, lib_count = row[:9]
        if (
            _clean_str(process_name) != process
            or _clean_str(thread_name) != thread
            or _clean_str(lib_name) != lib
        ):
            continue
        value = _to_int(lib_count)
        if value is None:
            continue
        total += value
        found = True
    return total, found


def _sum_function_sheet(
    ws: Any,
    *,
    process: str,
    thread: str,
    lib: str,
    function: str,
) -> tuple[int, bool]:
    """Return ``(sum_functionCount_total, found)`` for the matching function."""
    total = 0
    found = False
    for row in _iter_data_rows(ws, min_cols=13):
        (
            _pid, process_name, _pcnt, _tid, thread_name, _tcnt,
            _fid, lib_name, _lcnt, _funcid, function_name,
            _fc_self, fc_total,
        ) = row[:13]
        if (
            _clean_str(process_name) != process
            or _clean_str(thread_name) != thread
            or _clean_str(lib_name) != lib
            or _clean_str(function_name) != function
        ):
            continue
        value = _to_int(fc_total)
        if value is None:
            continue
        total += value
        found = True
    return total, found


# ---------------------------------------------------------------------------
# Workbook extraction for a specific target
# ---------------------------------------------------------------------------

LEVEL_CHOICES = ("total", "process", "thread", "lib", "function")

# Which name flags each level needs.  The flags above the chosen level
# are required because they identify *where* the target lives in the
# hierarchy, but levels below it are irrelevant.
_LEVEL_REQUIRED_NAMES: dict[str, tuple[str, ...]] = {
    "total": (),
    "process": ("process",),
    "thread": ("process", "thread"),
    "lib": ("process", "thread", "lib"),
    "function": ("process", "thread", "lib", "function"),
}


def _ensure_openpyxl() -> None:
    if load_workbook is None:
        raise RuntimeError(
            "openpyxl is required for report_compare.py. "
            "Install it on the Windows host with: pip install openpyxl"
        )


def _validate_target(
    level: str,
    *,
    process: str | None,
    thread: str | None,
    lib: str | None,
    function: str | None,
) -> None:
    if level not in LEVEL_CHOICES:
        raise ValueError(f"invalid level {level!r}; choose from {LEVEL_CHOICES}")
    required = _LEVEL_REQUIRED_NAMES[level]
    supplied = {
        "process": process,
        "thread": thread,
        "lib": lib,
        "function": function,
    }
    missing = [name for name in required if not supplied[name]]
    if missing:
        flag_list = ", ".join(f"--{name}" for name in missing)
        raise ValueError(f"--level {level} requires {flag_list}")


def extract_target_count(
    path: str,
    *,
    level: str,
    process: str | None = None,
    thread: str | None = None,
    lib: str | None = None,
    function: str | None = None,
) -> dict[str, Any]:
    """Open *path* and pull the count for the specified target.

    Returns a dict with:
      total:  total instruction count from page 1 (always populated
              when page 1 is present; useful context regardless of the
              requested level)
      value:  the target's count (summed across matching rows)
      found:  True if the target name was present at least once
      level:  echo of the requested level
    """
    _ensure_openpyxl()
    _validate_target(level, process=process, thread=thread, lib=lib, function=function)

    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        total: int | None = None
        processes: dict[str, int] = {}
        if PROCESS_SHEET in wb.sheetnames:
            total, processes = _parse_process_sheet(wb[PROCESS_SHEET])

        value = 0
        found = False

        if level == "total":
            if total is not None:
                value = total
                found = True

        elif level == "process":
            assert process is not None  # guaranteed by _validate_target
            process_value = processes.get(process)
            if process_value is not None:
                value = process_value
                found = True

        elif level == "thread":
            if THREAD_SHEET in wb.sheetnames:
                assert process is not None and thread is not None
                value, found = _sum_thread_sheet(wb[THREAD_SHEET], process=process, thread=thread)

        elif level == "lib":
            if LIB_SHEET in wb.sheetnames:
                assert process is not None and thread is not None and lib is not None
                value, found = _sum_lib_sheet(
                    wb[LIB_SHEET], process=process, thread=thread, lib=lib,
                )

        elif level == "function":
            if FUNCTION_SHEET in wb.sheetnames:
                assert (
                    process is not None and thread is not None
                    and lib is not None and function is not None
                )
                value, found = _sum_function_sheet(
                    wb[FUNCTION_SHEET],
                    process=process,
                    thread=thread,
                    lib=lib,
                    function=function,
                )
    finally:
        wb.close()

    return {
        "path": path,
        "level": level,
        "total": total,
        "value": value,
        "found": found,
    }


# ---------------------------------------------------------------------------
# Top-level compare
# ---------------------------------------------------------------------------

def _pct(delta: int, baseline: int) -> float | None:
    if not baseline:
        return None
    return round(delta / baseline * 100, 4)


def _pair_key(report: dict[str, Any]) -> tuple[str, int, int]:
    return (report["case"], report["round"], report["step"])


def _build_target_dict(
    level: str,
    *,
    process: str | None,
    thread: str | None,
    lib: str | None,
    function: str | None,
) -> dict[str, str | None]:
    """Return the subset of target names relevant to the chosen level."""
    required = _LEVEL_REQUIRED_NAMES[level]
    lookup = {"process": process, "thread": thread, "lib": lib, "function": function}
    return {name: lookup[name] for name in required}


def compare_reports(
    *,
    baseline_dir: str,
    candidate_dir: str,
    level: str = "total",
    process: str | None = None,
    thread: str | None = None,
    lib: str | None = None,
    function: str | None = None,
) -> dict[str, Any]:
    """Compare two report directories at the requested granularity.

    Pairs workbooks by ``(case, round, step)``.  For each pair extracts
    the count for the named target and reports baseline / candidate /
    delta.  Summed across pairs the same way — so one top-level number
    can be surfaced upstream.
    """
    _validate_target(level, process=process, thread=thread, lib=lib, function=function)

    baseline_reports = find_reports(baseline_dir)
    candidate_reports = find_reports(candidate_dir)
    b_map = {_pair_key(r): r for r in baseline_reports}
    c_map = {_pair_key(r): r for r in candidate_reports}
    all_keys = sorted(set(b_map) | set(c_map))

    pair_results: list[dict[str, Any]] = []
    agg_b = 0
    agg_c = 0
    agg_b_found = False
    agg_c_found = False
    pair_errors = False

    for key in all_keys:
        entry: dict[str, Any] = {
            "case": key[0],
            "round": key[1],
            "step": key[2],
        }
        b = b_map.get(key)
        c = c_map.get(key)

        if b is None or c is None:
            entry["missing"] = "baseline" if b is None else "candidate"
            entry["baseline_path"] = b["path"] if b else None
            entry["candidate_path"] = c["path"] if c else None
            pair_errors = True
            pair_results.append(entry)
            continue

        entry["baseline_path"] = b["path"]
        entry["candidate_path"] = c["path"]

        try:
            b_data = extract_target_count(
                b["path"], level=level,
                process=process, thread=thread, lib=lib, function=function,
            )
            c_data = extract_target_count(
                c["path"], level=level,
                process=process, thread=thread, lib=lib, function=function,
            )
        except Exception as exc:  # noqa: BLE001 — surface parse errors per-pair
            entry["error"] = f"failed to parse xlsx: {exc}"
            pair_errors = True
            pair_results.append(entry)
            continue

        b_val = b_data["value"]
        c_val = c_data["value"]
        entry["baseline"] = b_val
        entry["candidate"] = c_val
        entry["baseline_found"] = b_data["found"]
        entry["candidate_found"] = c_data["found"]
        entry["baseline_total"] = b_data["total"]
        entry["candidate_total"] = c_data["total"]
        delta = c_val - b_val
        entry["delta"] = delta
        entry["delta_pct"] = _pct(delta, b_val)

        agg_b += b_val
        agg_c += c_val
        if b_data["found"]:
            agg_b_found = True
        if c_data["found"]:
            agg_c_found = True

        pair_results.append(entry)

    agg_delta = agg_c - agg_b
    aggregate = {
        "baseline": agg_b,
        "candidate": agg_c,
        "delta": agg_delta,
        "delta_pct": _pct(agg_delta, agg_b),
        "baseline_found": agg_b_found,
        "candidate_found": agg_c_found,
        "pairs_compared": sum(1 for r in pair_results if "missing" not in r and "error" not in r),
        "pairs_missing_baseline": sum(1 for r in pair_results if r.get("missing") == "baseline"),
        "pairs_missing_candidate": sum(1 for r in pair_results if r.get("missing") == "candidate"),
    }

    return {
        "success": not pair_errors and bool(pair_results),
        "baseline_dir": os.path.abspath(baseline_dir),
        "candidate_dir": os.path.abspath(candidate_dir),
        "level": level,
        "target": _build_target_dict(
            level, process=process, thread=thread, lib=lib, function=function,
        ),
        "aggregate": aggregate,
        "reports": pair_results,
    }


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Compare two modelCase test report directories on instruction count.",
    )
    parser.add_argument("--baseline", required=True, help="Baseline (stock) report directory.")
    parser.add_argument("--candidate", required=True, help="Candidate (feature) report directory.")
    parser.add_argument(
        "--level",
        choices=LEVEL_CHOICES,
        default="total",
        help=(
            "Granularity of comparison (default: total). "
            "Each level requires the names at and above it: "
            "process -> --process; thread -> --process --thread; "
            "lib -> --process --thread --lib; function -> all four."
        ),
    )
    parser.add_argument("--process", default=None, help="Target processName.")
    parser.add_argument("--thread", default=None, help="Target threadName.")
    parser.add_argument("--lib", default=None, help="Target libName.")
    parser.add_argument("--function", default=None, help="Target functionName.")
    return parser


def main(argv: list[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    try:
        result = compare_reports(
            baseline_dir=args.baseline,
            candidate_dir=args.candidate,
            level=args.level,
            process=args.process,
            thread=args.thread,
            lib=args.lib,
            function=args.function,
        )
    except (RuntimeError, ValueError) as exc:
        json.dump({"success": False, "error": str(exc)}, sys.stdout, ensure_ascii=False, indent=2)
        sys.stdout.write("\n")
        return 1
    json.dump(result, sys.stdout, ensure_ascii=False, indent=2)
    sys.stdout.write("\n")
    return 0 if result.get("success") else 2


if __name__ == "__main__":
    raise SystemExit(main())
